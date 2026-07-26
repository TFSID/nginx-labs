"""Generate COMPARISON.xlsx and COMPARISON.html from repo data."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import json

OUT = Path(r"D:\TFS\Github\Project Infradyne\mws\projects\APL\APL_VAPT\nginx-poc")

REPOS = [
    ["DepthFirstDisclosures__Nginx-Rift", "Exploit+Lab", "Python,Shell", "🔴 High", 325, 7, "Yes", "Yes", "No", "Original discoverer, heap spray feng shui"],
    ["0xBlackash__CVE-2026-42945", "Exploit+Lab", "Python,Shell", "🔴 High", 325, 8, "Yes", "Yes", "No", "Clone of DepthFirst"],
    ["bamov970__CVE-2026-42945-Nginx-RCE-bypass-ASLR", "Exploit+Scan", "Python", "🔴 Very High", 800, 12, "Yes", "No", "Yes", "Satu-satunya ASLR bypass via /proc/self/mem"],
    ["cipherspy__CVE-2026-42945-POC", "Exploit+Helper", "Python,Shell", "🟡 Medium", 300, 4, "Yes", "No", "No", "HTB-focused, cheat sheet, helpers"],
    ["dinosn__cve-2026-42945-nginx32-lab", "Exploit+Lab", "Python,Shell", "🔴 High", 400, 12, "Yes", "Yes", "N/A (32-bit)", "32-bit specific, remote bruteforce"],
    ["F2u0a0d3__CVE-2026-42945-nginx-rift-poc", "Exploit+Research", "Python,Shell", "🔴 Very High", 600, 16, "Yes", "Yes", "Partial", "Brute-sweep, address discovery, vuln vs patched"],
    ["gagaltotal__CVE-2026-42945-NGINX-Rift-Toolkit", "Toolkit", "Python", "🔴 Very High", 2250, 7, "Yes", "No", "No", "All-in-one: scan, exploit, patch, HTML reports"],
    ["MateusVerass__nGixshell", "Exploit+WebAudit", "Python,Shell", "🔴 Very High", 2000, 7, "Yes", "No", "No", "Full web audit, WAF bypass, CVE DB, subdomain scan"],
    ["rheodev__CVE-2026-42945", "Exploit+Ref", "Python,C", "🟡 Medium", 250, 4, "Yes", "No", "No", "DoS PoC + C analysis code"],
    ["josephfelix__CVE-2026-42945-nginx-rift", "Exploit+Lab", "Python,Shell", "🔴 High", 325, 7, "Yes", "Yes", "No", "Fork of DepthFirst"],
    ["nu0l__NGINX-Rift", "Scanner", "Go", "🔴 Very High", 700, 4, "No", "No", "No", "Satu-satunya Go, dual-CVE (42945+9256), K8s-aware"],
    ["soksofos__wazuh-nginx-cve-2026-42945-sca-lab", "Wazuh", "YAML", "🟡 Medium", 400, 10, "No", "No", "No", "Satu-satunya Wazuh SCA policy"],
    ["BarAppTeam__nginx-cve-fix", "Fix/Patch", "Shell,Makefile", "🔴 Very High", 850, 18, "No (fix)", "Yes", "N/A", "Production-grade patching, VEX, SBOM"],
    ["strivepan__Nginx_cve-2026-42945-scanner-gui", "GUI Tool", "Electron", "🟡 Medium", 0, 1, "No", "No", "No", "Satu-satunya GUI tool (Electron)"],
    ["friparia__NGINX_RIFT_SCAN_CVE_2026_42945", "Scanner", "Python", "🟢 Low", 150, 4, "No", "No", "No", "Bilingual (FR/EN) README"],
    ["iammerrida-source__nginx-rift-detect", "Scanner", "Python", "🟢 Low", 100, 2, "No", "No", "No", "Lightweight detection"],
    ["limo57640-crypto__nginx-rift-detector", "Scanner", "Python", "🟢 Low", 100, 2, "No", "No", "No", "Basic version detector"],
    ["realityone__cve-2026-42945-scan", "Scanner", "Python", "🟢 Low", 150, 3, "No", "No", "No", "Multi-host scanning"],
    ["simota__nginx-rift-scanner", "Scanner", "Python", "🟢 Low", 100, 2, "No", "No", "No", "Minimal scanner"],
    ["oseasfr__CVE_2026-42945", "Scanner", "Python", "🟢 Low", 80, 2, "No", "No", "No", "Minimal scanner"],
    ["sibersan__web-server-audit_CVE-2026-42945", "Scanner/Audit", "Python", "🟡 Medium", 200, 3, "No", "No", "No", "Broader web audit"],
    ["chenqin231__CVE-2026-42945", "Scanner+Bash", "Shell", "🟢 Low", 80, 2, "No", "No", "No", "Pure Bash scanner"],
    ["hnytgl__CVE-2026-42945", "Reference", "None", "⚪ Very Low", 0, 1, "No", "No", "No", "Info only"],
    ["tal7aouy__nginx-cve-2026-42945", "Reference", "None", "⚪ Very Low", 0, 1, "No", "No", "No", "Clean informational page"],
    ["LiaoZiqi-GZFLS__CVE-2026-42945", "Reference", "Shell,Python", "🟢 Low", 120, 3, "No", "No", "No", "Chinese language analysis"],
    ["fkj-src__fix_nginx_cve_2026_42945", "Fix/Patch", "Shell", "🟢 Low", 30, 2, "No (fix)", "No", "N/A", "Minimalist 1-line patch"],
]

HEADERS = ["#", "Repo Name", "Type", "Language", "Complexity", "LOC", "Files", "Working Exploit", "Docker", "ASLR Bypass", "Fitur Unik"]

# ============================================================================
# XLSX
# ============================================================================
def gen_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Full Comparison"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="center", wrap_text=True)
    thin = Side(style="thin", color="B4C6E7")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    alt_fill = PatternFill("solid", fgColor="D6E4F0")
    red_fill = PatternFill("solid", fgColor="FFF2CC")  # high complexity

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 7
    ws.column_dimensions["G"].width = 7
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 9
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 50

    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = border

    for ri, repo in enumerate(REPOS, 2):
        row = [ri - 1] + repo
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = cell_align
            c.border = border
            if (ri - 2) % 2 == 1:
                c.fill = alt_fill
        # highlight complex repos
        if "Very High" in str(repo[3]):
            ws.cell(row=ri, column=5).fill = red_fill

    # --- Sheet 2: Scoring ---
    ws2 = wb.create_sheet("Scoring Matrix")
    score_headers = ["Repo Name", "Effectiveness (1-10)", "Complexity (1-10)", "RCE Exploit", "ASLR Bypass", "Docker Ready", "Multi-Fitur", "Production Grade"]
    for ci, h in enumerate(score_headers, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = border

    scores = [
        ["bamov970", 9, 9, 4, 3, 0, 2, 0],
        ["gagaltotal", 7, 10, 4, 0, 0, 2, 1],
        ["F2u0a0d3", 9, 8, 4, 1, 2, 2, 0],
        ["DepthFirst", 6, 5, 4, 0, 2, 0, 0],
        ["MateusVerass", 7, 10, 4, 0, 0, 2, 1],
        ["dinosn", 7, 5, 4, 0, 2, 1, 0],
        ["nu0l", 3, 6, 0, 0, 0, 2, 1],
        ["BarAppTeam", 4, 8, 0, 0, 2, 1, 1],
        ["cipherspy", 5, 4, 4, 0, 0, 1, 0],
    ]

    for ri, row in enumerate(scores, 2):
        for ci, val in enumerate(row, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border

    # conditional formatting for scores
    green = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red_fill2 = PatternFill("solid", fgColor="FFC7CE")
    for ri in range(2, 2 + len(scores)):
        c = ws2.cell(row=ri, column=2)  # effectiveness
        if isinstance(c.value, int):
            if c.value >= 8:
                c.fill = green
            elif c.value >= 5:
                c.fill = yellow
            else:
                c.fill = red_fill2

    ws2.column_dimensions["A"].width = 16
    for i in range(2, 9):
        ws2.column_dimensions[get_column_letter(i)].width = 18

    wb.save(str(OUT / "COMPARISON.xlsx"))
    print(f"Created COMPARISON.xlsx ({len(REPOS)} repos, 2 sheets)")

# ============================================================================
# HTML
# ============================================================================
def gen_html():
    rows_json = json.dumps([[i] + r for i, r in enumerate(REPOS, 1)], indent=0)

    html = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>CVE-2026-42945 Repository Comparison</title>
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: -apple-system, 'Segoe UI', Roboto, sans-serif; background: #0d1117; color: #e6edf3; padding: 2rem; }
h1 { color: #58a6ff; font-size: 1.8rem; margin-bottom: .3rem; }
.subtitle { color: #8b949e; margin-bottom: 1.5rem; font-size: .9rem; }
.controls { display: flex; gap: 1rem; flex-wrap: wrap; margin-bottom: 1.5rem; align-items: center; }
.controls input, .controls select { padding: .5rem .8rem; border: 1px solid #30363d; border-radius: 6px; background: #161b22; color: #e6edf3; font-size: .9rem; }
.controls input { flex: 1; min-width: 200px; }
.controls select { min-width: 140px; }
.controls label { color: #8b949e; font-size: .85rem; }
#counter { color: #8b949e; font-size: .85rem; margin-left: auto; }
table { width: 100%; border-collapse: collapse; font-size: .85rem; }
th { background: #161b22; color: #58a6ff; padding: .6rem .5rem; text-align: left; font-weight: 600; border-bottom: 2px solid #30363d; cursor: pointer; position: relative; user-select: none; white-space: nowrap; }
th:hover { background: #1c2128; }
th::after { content: ' \25B4\25BE'; font-size: .6rem; color: #484f58; margin-left: 3px; }
th.asc::after { content: ' \25B2'; color: #58a6ff; }
th.desc::after { content: ' \25BC'; color: #58a6ff; }
td { padding: .5rem; border-bottom: 1px solid #21262d; vertical-align: top; }
tr:hover td { background: #161b22; }
.complexity-High { color: #f85149; }
.complexity-Very-High { color: #f85149; font-weight: 700; }
.complexity-Medium { color: #d29922; }
.complexity-Low { color: #3fb950; }
.complexity-Very-Low { color: #8b949e; }
.badge { display: inline-block; padding: .1rem .4rem; border-radius: 3px; font-size: .75rem; font-weight: 600; }
.badge-yes { background: #1b4a23; color: #3fb950; }
.badge-no { background: #3d1f1f; color: #f85149; }
.badge-partial { background: #3d2f00; color: #d29922; }
.badge-na { background: #1c2128; color: #8b949e; }
.tag { display: inline-block; padding: .1rem .4rem; border-radius: 3px; background: #1c2128; color: #8b949e; font-size: .75rem; margin: 1px; }
.exploit-row { border-left: 3px solid #3fb950; }
.scanner-row { border-left: 3px solid #d29922; }
.fix-row { border-left: 3px solid #58a6ff; }
.reference-row { border-left: 3px solid #484f58; }
.toolkit-row { border-left: 3px solid #bc8cff; }
@media (max-width: 768px) {
  body { padding: 1rem; }
  .controls { flex-direction: column; }
  .controls input { width: 100%; }
  #counter { margin-left: 0; }
  th, td { padding: .4rem .3rem; font-size: .78rem; }
}
footer { margin-top: 2rem; padding-top: 1rem; border-top: 1px solid #21262d; color: #484f58; font-size: .8rem; text-align: center; }
</style>
</head>
<body>
<h1>CVE-2026-42945 (NGINX Rift) — Repository Comparison</h1>
<p class="subtitle">Total: <span id="totalCount">0</span> repositories · Last updated: 2026-06-13</p>

<div class="controls">
  <input type="text" id="search" placeholder="Search repos, features, languages..." oninput="filterTable()">
  <div>
    <label for="typeFilter">Type: </label>
    <select id="typeFilter" onchange="filterTable()">
      <option value="">All Types</option>
      <option value="Exploit">Exploit</option>
      <option value="Scanner">Scanner</option>
      <option value="Toolkit">Toolkit</option>
      <option value="Fix/Patch">Fix/Patch</option>
      <option value="Reference">Reference</option>
      <option value="Wazuh">Wazuh</option>
      <option value="GUI">GUI</option>
    </select>
  </div>
  <div>
    <label for="exploitFilter">Exploit: </label>
    <select id="exploitFilter" onchange="filterTable()">
      <option value="">All</option>
      <option value="Yes">Working Exploit</option>
      <option value="No">No Exploit</option>
    </select>
  </div>
  <div>
    <label for="dockerFilter">Docker: </label>
    <select id="dockerFilter" onchange="filterTable()">
      <option value="">All</option>
      <option value="Yes">Has Docker</option>
      <option value="No">No Docker</option>
    </select>
  </div>
  <span id="counter">0 repos shown</span>
</div>

<table>
<thead>
<tr>
  <th onclick="sortTable(0)">#</th>
  <th onclick="sortTable(1)" class="asc">Repo Name</th>
  <th onclick="sortTable(2)">Type</th>
  <th onclick="sortTable(3)">Language</th>
  <th onclick="sortTable(4)">Complexity</th>
  <th onclick="sortTable(5)">LOC</th>
  <th onclick="sortTable(6)">Files</th>
  <th onclick="sortTable(7)">Working Exploit</th>
  <th onclick="sortTable(8)">Docker</th>
  <th onclick="sortTable(9)">ASLR Bypass</th>
  <th onclick="sortTable(10)">Fitur Unik</th>
</tr>
</thead>
<tbody id="tableBody"></tbody>
</table>

<footer>
  Generated by MWS Security Research · PT Meta Wangsa Solusi
</footer>

<script>
const data = """ + rows_json + r""";

function badge(val) {
  if (!val) return '<span class="badge badge-no">No</span>';
  const s = String(val).toLowerCase();
  if (s === 'yes' || s.startsWith('yes')) return '<span class="badge badge-yes">Yes</span>';
  if (s.startsWith('no')) return '<span class="badge badge-no">' + val + '</span>';
  if (s.startsWith('partial') || s.startsWith('n/a') || s.includes('32')) return '<span class="badge badge-partial">' + val + '</span>';
  return '<span class="badge badge-na">' + val + '</span>';
}

function complexityClass(c) {
  const s = c.replace(/[^a-zA-Z-]/g, '');
  return 'complexity-' + s;
}

function typeFilterClass(type) {
  if (type.startsWith('Exploit')) return 'exploit-row';
  if (type.startsWith('Scanner')) return 'scanner-row';
  if (type.startsWith('Fix') || type.startsWith('Patch')) return 'fix-row';
  if (type.startsWith('Reference')) return 'reference-row';
  if (type === 'Toolkit') return 'toolkit-row';
  return '';
}

function render(rows) {
  const tbody = document.getElementById('tableBody');
  tbody.innerHTML = rows.map(r => {
    const c = r[4];
    return '<tr class="' + typeFilterClass(r[2]) + '">'
      + '<td>' + r[0] + '</td>'
      + '<td><code style="color:#e6edf3;font-size:.82rem">' + r[1] + '</code></td>'
      + '<td>' + r[2] + '</td>'
      + '<td>' + (r[3]||'').split(',').map(t => '<span class="tag">' + t.trim() + '</span>').join('') + '</td>'
      + '<td class="' + complexityClass(c) + '">' + c + '</td>'
      + '<td>' + r[5] + '</td>'
      + '<td>' + r[6] + '</td>'
      + '<td>' + badge(r[7]) + '</td>'
      + '<td>' + badge(r[8]) + '</td>'
      + '<td>' + badge(r[9]) + '</td>'
      + '<td style="color:#8b949e;font-size:.8rem">' + r[10] + '</td>'
      + '</tr>';
  }).join('');
  document.getElementById('totalCount').textContent = data.length;
  document.getElementById('counter').textContent = rows.length + ' repos shown';
}

function filterTable() {
  const q = document.getElementById('search').value.toLowerCase();
  const typeFilter = document.getElementById('typeFilter').value.toLowerCase();
  const exploitFilter = document.getElementById('exploitFilter').value.toLowerCase();
  const dockerFilter = document.getElementById('dockerFilter').value.toLowerCase();

  let filtered = data.filter(r => {
    const rowText = r.join(' ').toLowerCase();
    if (q && !rowText.includes(q)) return false;
    if (typeFilter && !r[2].toLowerCase().includes(typeFilter)) return false;
    if (exploitFilter && !r[7].toLowerCase().startsWith(exploitFilter)) return false;
    if (dockerFilter && !r[8].toLowerCase().startsWith(dockerFilter)) return false;
    return true;
  });
  render(filtered);
}

let sortDir = {};
function sortTable(col) {
  sortDir[col] = sortDir[col] === 'asc' ? 'desc' : 'asc';
  const dir = sortDir[col] === 'asc' ? 1 : -1;
  data.sort((a, b) => {
    let va = a[col], vb = b[col];
    if (typeof va === 'string') va = va.toLowerCase();
    if (typeof vb === 'string') vb = vb.toLowerCase();
    if (!isNaN(va) && !isNaN(vb)) return (va - vb) * dir;
    return String(va).localeCompare(String(vb)) * dir;
  });

  // update header arrows
  document.querySelectorAll('th').forEach(th => th.classList.remove('asc', 'desc'));
  event.target.classList.add(dir === 1 ? 'asc' : 'desc');
  filterTable();
}

filterTable();
</script>
</body>
</html>"""

    (OUT / "COMPARISON.html").write_text(html)
    print(f"Created COMPARISON.html ({len(html)} bytes)")


if __name__ == "__main__":
    gen_xlsx()
    gen_html()
    print("Done!")
